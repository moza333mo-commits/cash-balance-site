from flask import Flask, jsonify, render_template
from openpyxl import load_workbook
import os

app = Flask(__name__)

def get_excel_value():
    file_path = 'data_live.xlsx'
    try:
        if not os.path.exists(file_path):
            return None, "ملف الإكسيل (data_live.xlsx) غير موجود"

        wb = load_workbook(file_path, data_only=True)
        sheet = wb.active

        # الصف الثالث (index = 3 لأن openpyxl يبدأ من 1)
        row_3 = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))[0]

        val = None

        # العمود F = index 5
        if len(row_3) > 5:
            val = row_3[5]

        # إذا الخلية فاضية → دور على أول رقم في الصف
        if val is None or str(val).strip() == '':
            for item in row_3:
                if isinstance(item, (int, float)):
                    return float(item), None
            return None, "الخلية فارغة ولا يوجد أي رقم في الصف"

        return float(val), None

    except Exception as e:
        return None, f"خطأ في قراءة الملف: {str(e)}"


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/data')
def data():
    val, error = get_excel_value()

    if error:
        response = jsonify({"error": error})
    else:
        response = jsonify({"value": val})

    # منع الكاش
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"

    return response


if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)
